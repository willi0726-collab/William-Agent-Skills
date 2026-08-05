#!/usr/bin/env python
"""Analyze Amazon Ads CSV/XLSX exports and write optimization recommendations."""

from __future__ import annotations

import argparse
import csv
import math
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional


COLUMN_ALIASES = {
    "campaign": ["campaign", "campaign name", "广告活动", "广告活动名称"],
    "ad_group": ["ad group", "ad group name", "ad group name (group)", "广告组", "广告组名称"],
    "keyword": ["keyword", "keyword text", "targeting", "target", "关键词", "投放", "投放词", "定向"],
    "search_term": ["customer search term", "search term", "matched search term", "搜索词", "买家搜索词", "客户搜索词"],
    "match_type": ["match type", "匹配类型"],
    "impressions": ["impressions", "impr.", "展示量", "曝光量"],
    "clicks": ["clicks", "点击量", "点击次数"],
    "spend": ["spend", "cost", "广告花费", "花费", "成本"],
    "sales": ["sales", "7 day total sales", "14 day total sales", "attributed sales", "销售额", "广告销售额"],
    "orders": ["orders", "purchases", "7 day total orders", "14 day total orders", "conversions", "订单", "订单量", "转化"],
    "acos": ["acos", "advertising cost of sales", "广告销售成本"],
    "cpc": ["cpc", "cost per click", "每次点击费用", "平均点击成本"],
    "bid": ["bid", "keyword bid", "default bid", "竞价", "出价"],
    "budget": ["budget", "daily budget", "预算", "日预算"],
}


def normalize_header(value: str) -> str:
    return re.sub(r"\s+", " ", str(value).strip().lower().replace("\ufeff", ""))


def parse_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text or text in {"-", "--", "N/A", "n/a"}:
        return None
    is_percent = "%" in text
    text = re.sub(r"[\$,￥¥,%\s]", "", text)
    text = text.replace(",", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return number / 100 if is_percent else number


def read_csv(path: Path) -> List[Dict[str, Any]]:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return list(csv.DictReader(handle))
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Could not decode CSV: {path}")


def read_xlsx(path: Path) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Reading XLSX requires openpyxl. Install it or export CSV.") from exc
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell or "").strip() for cell in rows[0]]
    records = []
    for row in rows[1:]:
        if not any(cell not in (None, "") for cell in row):
            continue
        records.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return records


def read_rows(path: Path) -> List[Dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx(path)
    raise ValueError("Input must be .csv, .xlsx, or .xlsm")


def build_column_map(headers: Iterable[str]) -> Dict[str, str]:
    normalized = {normalize_header(header): header for header in headers}
    mapping = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            found = normalized.get(normalize_header(alias))
            if found:
                mapping[canonical] = found
                break
    return mapping


def get_text(row: Dict[str, Any], mapping: Dict[str, str], key: str) -> str:
    source = mapping.get(key)
    value = row.get(source) if source else ""
    return "" if value is None else str(value).strip()


def get_num(row: Dict[str, Any], mapping: Dict[str, str], key: str) -> Optional[float]:
    source = mapping.get(key)
    return parse_number(row.get(source)) if source else None


def fmt_money(value: Optional[float]) -> str:
    return "" if value is None else f"{value:.2f}"


def fmt_pct(value: Optional[float]) -> str:
    return "" if value is None else f"{value:.1%}"


def priority_for(action: str) -> str:
    if action in {"negative_exact", "budget_up"}:
        return "high"
    if action in {"bid_down", "harvest_keyword", "harvest_asin"}:
        return "medium"
    return "low"


def recommendation_rows(rows: List[Dict[str, Any]], target_acos: float, no_sale_clicks: int, min_harvest_orders: int) -> List[Dict[str, str]]:
    if not rows:
        return []
    mapping = build_column_map(rows[0].keys())
    output: List[Dict[str, str]] = []

    for index, row in enumerate(rows, start=2):
        campaign = get_text(row, mapping, "campaign")
        ad_group = get_text(row, mapping, "ad_group")
        keyword = get_text(row, mapping, "keyword")
        search_term = get_text(row, mapping, "search_term")
        target_label = search_term or keyword

        impressions = get_num(row, mapping, "impressions") or 0
        clicks = get_num(row, mapping, "clicks") or 0
        spend = get_num(row, mapping, "spend") or 0
        sales = get_num(row, mapping, "sales") or 0
        orders = get_num(row, mapping, "orders") or 0
        bid = get_num(row, mapping, "bid")
        acos = get_num(row, mapping, "acos")
        cpc = get_num(row, mapping, "cpc")

        if acos is None and sales > 0:
            acos = spend / sales
        if cpc is None and clicks > 0:
            cpc = spend / clicks
        ctr = clicks / impressions if impressions > 0 else None
        cvr = orders / clicks if clicks > 0 else None

        actions = []
        if clicks >= no_sale_clicks and orders == 0 and spend > 0 and search_term:
            actions.append(("negative_exact", None, f"Add exact negative for '{search_term}' or reduce bid if strategically important."))
        elif clicks >= max(8, int(no_sale_clicks * 0.7)) and orders == 0 and spend > 0:
            new_bid = bid * 0.75 if bid else None
            actions.append(("bid_down", new_bid, "No orders after meaningful clicks; reduce bid 15-30% or isolate for testing."))

        if orders > 0 and acos is not None and acos > target_acos * 1.3:
            cut = 0.8 if acos <= target_acos * 2 else 0.7
            new_bid = bid * cut if bid else None
            actions.append(("bid_down", new_bid, f"ACOS {fmt_pct(acos)} is above target {fmt_pct(target_acos)}."))

        if orders >= min_harvest_orders and acos is not None and acos <= target_acos * 0.7:
            new_bid = bid * 1.10 if bid else None
            actions.append(("bid_up", new_bid, f"Efficient converter below target ACOS; consider cautious bid increase."))

        if search_term and orders >= min_harvest_orders and acos is not None and acos <= target_acos:
            action = "harvest_asin" if re.search(r"\b(b0[a-z0-9]{8}|asin)\b", search_term, re.I) else "harvest_keyword"
            actions.append((action, None, f"Harvest '{search_term}' into exact targeting with controlled bid."))

        if not actions and clicks > 0 and (acos is None or sales == 0) and spend > 0:
            actions.append(("investigate", None, "Review query relevance, listing conversion, and attribution before changing."))

        for action, suggested_bid, recommendation in actions:
            output.append({
                "priority": priority_for(action),
                "action": action,
                "source_row": str(index),
                "campaign": campaign,
                "ad_group": ad_group,
                "keyword_or_target": keyword,
                "search_term": search_term,
                "clicks": f"{clicks:.0f}",
                "spend": fmt_money(spend),
                "sales": fmt_money(sales),
                "orders": f"{orders:.0f}",
                "acos": fmt_pct(acos),
                "ctr": fmt_pct(ctr),
                "cvr": fmt_pct(cvr),
                "cpc": fmt_money(cpc),
                "current_bid": fmt_money(bid),
                "suggested_bid": fmt_money(suggested_bid),
                "recommendation": recommendation,
                "risk_note": "Review brand terms, launch goals, margin, inventory, Buy Box, and seasonality before applying.",
            })

    return output


def write_csv(path: Path, rows: List[Dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "priority", "action", "source_row", "campaign", "ad_group", "keyword_or_target", "search_term",
        "clicks", "spend", "sales", "orders", "acos", "ctr", "cvr", "cpc", "current_bid", "suggested_bid",
        "recommendation", "risk_note",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, help="Amazon Ads CSV/XLSX export")
    parser.add_argument("--output", required=True, help="Recommendations CSV path")
    parser.add_argument("--target-acos", type=float, default=0.30, help="Target ACOS as decimal, e.g. 0.30")
    parser.add_argument("--no-sale-clicks", type=int, default=12, help="Clicks with zero orders before negative/bid-down recommendation")
    parser.add_argument("--min-harvest-orders", type=int, default=2, help="Minimum orders before harvesting a search term")
    args = parser.parse_args()

    rows = read_rows(Path(args.input))
    recommendations = recommendation_rows(rows, args.target_acos, args.no_sale_clicks, args.min_harvest_orders)
    write_csv(Path(args.output), recommendations)
    print(f"Read {len(rows)} rows; wrote {len(recommendations)} recommendations to {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
