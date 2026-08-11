import importlib.util
import tempfile
import unittest
from pathlib import Path

import openpyxl


SKILL_DIR = Path(__file__).parent
SCRIPT = SKILL_DIR / "fill_phone_case_template.py"
REQUIRED_HEADERS = [
    "feed_product_type", "item_sku", "brand_name", "parent_child",
    "variation_theme", "item_name", "manufacturer", "update_delete",
    "product_description", "external_product_id_type", "parent_sku",
    "relationship_type", "bullet_point1", "bullet_point2", "bullet_point3",
    "bullet_point4", "bullet_point5", "generic_keywords", "special_features1",
    "size_name", "size_map", "color_name", "color_map", "material_type",
    "pattern_name", "theme", "form_factor", "included_components",
    "compatible_phone_models1", "package_height", "package_length",
    "package_width", "package_weight", "package_height_unit_of_measure",
    "package_length_unit_of_measure", "package_width_unit_of_measure",
    "package_weight_unit_of_measure", "country_of_origin",
    "fulfillment_availability#1.quantity",
    "purchasable_offer[marketplace_id=ATVPDKIKX0DER]#1.our_price#1.schedule#1.value_with_tax",
    "list_price", "condition_type",
]


class FillTemplateTest(unittest.TestCase):
    def test_every_color_model_child_gets_commercial_fields(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            template = tmp / "template.xlsm"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Template"
            for column, header in enumerate(REQUIRED_HEADERS, 1):
                ws.cell(3, column, header)
            wb.save(template)

            spec = importlib.util.spec_from_file_location("phone_case_fill", SCRIPT)
            module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(module)
            module.INPUT_TEMPLATE = str(template)
            module.OUTPUT_DIR = str(tmp / "output")
            module.MODELS = [("IP16", "iPhone 16", "iPhone 16"), ("IP17", "iPhone 17", "iPhone 17")]
            module.COLORS = [("Pink", "Pink", "PINK", "Pink"), ("Blue", "Blue", "BLUE", "Blue")]
            module.VARIATION_THEME = "ColorName-SizeName"
            module.fill_template()

            output = Path(module.OUTPUT_DIR) / f"phone_case_{module.BRAND}_{module.PARENT_SKU}.xlsm"
            result = openpyxl.load_workbook(output, data_only=False)
            sheet = result["Template"]
            columns = {sheet.cell(3, c).value: c for c in range(1, sheet.max_column + 1)}
            for row in range(5, 9):
                self.assertEqual(module.PRICE, sheet.cell(row, columns["list_price"]).value)
                self.assertEqual("CN", sheet.cell(row, columns["country_of_origin"]).value)
                self.assertEqual("New", sheet.cell(row, columns["condition_type"]).value)
                self.assertEqual(module.VARIATION_THEME, sheet.cell(row, columns["variation_theme"]).value)
            result.close()

    def test_latest_template_uses_settings_rows_and_item_highlight(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            template = tmp / "latest.xlsm"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Template"
            ws.cell(1, 1, "settings=labelRow=4&attributeRow=5&dataRow=8")
            latest = [
                ("SKU", "item_sku"), ("Product Type", "product_type"),
                ("Listing Action", "record_action"),
                ("Parentage Level", "parentage_level[marketplace_id=ATVPDKIKX0DER]#1.value"),
                ("Parent SKU", "child_parent_sku_relationship[marketplace_id=ATVPDKIKX0DER]#1.parent_sku"),
                ("Variation Theme Name", "variation_theme#1.name"),
                ("Item Name", "item_name[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"),
                ("Item Highlight", "title_differentiation[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"),
                ("Brand Name", "brand[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"),
                ("Manufacturer", "manufacturer[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"),
                ("Product Description", "product_description[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"),
                ("Bullet Point", "bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"),
                ("Bullet Point", "bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#2.value"),
                ("Bullet Point", "bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#3.value"),
                ("Bullet Point", "bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#4.value"),
                ("Bullet Point", "bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#5.value"),
                ("Generic Keyword", "generic_keyword[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"),
                ("Special Features", "special_feature[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"),
                ("Color", "color[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"),
                ("Compatible Phone Models", "compatible_phone_models[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"),
                ("Material", "material[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"),
                ("Product Id Type", "amzn1.volt.ca.product_id_type"),
                ("Number of Items", "number_of_items[marketplace_id=ATVPDKIKX0DER]#1.value"),
                ("Part Number", "part_number[marketplace_id=ATVPDKIKX0DER]#1.value"),
                ("Compatible Devices", "compatible_devices[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"),
                ("Finish Type", "finish_type[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"),
                ("Item Length", "item_length[marketplace_id=ATVPDKIKX0DER]#1.value"),
                ("Item Length Unit", "item_length[marketplace_id=ATVPDKIKX0DER]#1.unit"),
                ("Item Dimension Length", "item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.value"),
                ("Item Dimension Length Unit", "item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.unit"),
                ("Item Dimension Width", "item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.value"),
                ("Item Dimension Width Unit", "item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.unit"),
                ("Item Dimension Height", "item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.value"),
                ("Item Dimension Height Unit", "item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.unit"),
                ("Are batteries required?", "batteries_required[marketplace_id=ATVPDKIKX0DER]#1.value"),
                ("Pattern", "pattern[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"),
                ("Theme", "theme[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"),
                ("Form Factor", "form_factor[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"),
                ("Skip Offer", "skip_offer[marketplace_id=ATVPDKIKX0DER]#1.value"),
                ("Item Condition", "condition_type[marketplace_id=ATVPDKIKX0DER]#1.value"),
                ("List Price", "list_price[marketplace_id=ATVPDKIKX0DER]#1.value"),
                ("Fulfillment Channel", "fulfillment_availability#1.fulfillment_channel_code"),
                ("Quantity", "fulfillment_availability#1.quantity"),
                ("Your Price", "purchasable_offer[marketplace_id=ATVPDKIKX0DER][audience=ALL]#1.our_price#1.schedule#1.value_with_tax"),
                ("Shipping Template", "merchant_shipping_group[marketplace_id=ATVPDKIKX0DER]#1.value"),
                ("Package Length", "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.value"),
                ("Package Length Unit", "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.unit"),
                ("Package Width", "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.value"),
                ("Package Width Unit", "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.unit"),
                ("Package Height", "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.value"),
                ("Package Height Unit", "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.unit"),
                ("Package Weight", "item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.value"),
                ("Package Weight Unit", "item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.unit"),
                ("Country of Origin", "country_of_origin[marketplace_id=ATVPDKIKX0DER]#1.value"),
            ]
            for column, (label, attribute) in enumerate(latest, 1):
                ws.cell(4, column, label)
                ws.cell(5, column, attribute)
            wb.save(template)

            spec = importlib.util.spec_from_file_location("phone_case_fill_latest", SCRIPT)
            module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(module)
            module.INPUT_TEMPLATE = str(template)
            module.OUTPUT_DIR = str(tmp / "output")
            module.MODELS = [("16", "iPhone 16", "iPhone 16")]
            module.COLORS = [("Black Lace", "Black", "BLACK", "Black Lace")]
            module.VARIATION_THEME = "COLOR/COMPATIBLE_PHONE_MODELS"
            module.ITEM_HIGHLIGHT = "Soft TPU lace case with detachable beaded wrist strap"
            module.SKIP_OFFER = False
            module.SHIPPING_TEMPLATE = "Migrated Template"
            module.FULFILLMENT_CHANNEL = "Fulfillment by Merchant (Default)"
            module.PRICE = 22.5
            module.QUANTITY = 0
            module.BULLET1 = module.BULLET2 = module.BULLET3 = module.BULLET4 = module.BULLET5 = "Supported product fact"
            module.SEARCH_TERMS = "lace phone case beaded wrist strap soft tpu"
            module.fill_template()

            output = Path(module.OUTPUT_DIR) / f"phone_case_{module.BRAND}_{module.PARENT_SKU}.xlsm"
            result = openpyxl.load_workbook(output, data_only=False)
            sheet = result["Template"]
            columns = {sheet.cell(5, c).value: c for c in range(1, sheet.max_column + 1)}
            self.assertEqual(module.PARENT_SKU, sheet.cell(8, columns["item_sku"]).value)
            self.assertEqual("Create or Replace (Full Update)", sheet.cell(8, columns["record_action"]).value)
            self.assertEqual("GTIN Exempt", sheet.cell(8, columns["amzn1.volt.ca.product_id_type"]).value)
            self.assertTrue(sheet.cell(8, columns["product_description[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"]).value)
            self.assertTrue(sheet.cell(8, columns["bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"]).value)
            self.assertEqual("CN", sheet.cell(8, columns["country_of_origin[marketplace_id=ATVPDKIKX0DER]#1.value"]).value)
            self.assertEqual("No", sheet.cell(8, columns["batteries_required[marketplace_id=ATVPDKIKX0DER]#1.value"]).value)
            self.assertEqual(module.ITEM_HIGHLIGHT, sheet.cell(9, columns["title_differentiation[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"]).value)
            self.assertIsNone(sheet.cell(9, columns["skip_offer[marketplace_id=ATVPDKIKX0DER]#1.value"]).value)
            self.assertEqual("Migrated Template", sheet.cell(9, columns["merchant_shipping_group[marketplace_id=ATVPDKIKX0DER]#1.value"]).value)
            self.assertEqual("Fulfillment by Merchant (Default)", sheet.cell(9, columns["fulfillment_availability#1.fulfillment_channel_code"]).value)
            self.assertEqual(0, sheet.cell(9, columns["fulfillment_availability#1.quantity"]).value)
            self.assertEqual(22.5, sheet.cell(9, columns["purchasable_offer[marketplace_id=ATVPDKIKX0DER][audience=ALL]#1.our_price#1.schedule#1.value_with_tax"]).value)
            self.assertEqual("GTIN Exempt", sheet.cell(9, columns["amzn1.volt.ca.product_id_type"]).value)
            self.assertEqual(1, sheet.cell(9, columns["number_of_items[marketplace_id=ATVPDKIKX0DER]#1.value"]).value)
            self.assertEqual(sheet.cell(9, columns["item_sku"]).value, sheet.cell(9, columns["part_number[marketplace_id=ATVPDKIKX0DER]#1.value"]).value)
            self.assertEqual("iPhone 16", sheet.cell(9, columns["compatible_devices[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"]).value)
            self.assertEqual("Matte", sheet.cell(9, columns["finish_type[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"]).value)
            self.assertEqual(4.72, sheet.cell(9, columns["item_length[marketplace_id=ATVPDKIKX0DER]#1.value"]).value)
            self.assertEqual("Inches", sheet.cell(9, columns["item_length[marketplace_id=ATVPDKIKX0DER]#1.unit"]).value)
            self.assertEqual(12, sheet.cell(9, columns["item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.value"]).value)
            self.assertEqual(5, sheet.cell(9, columns["item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.value"]).value)
            self.assertEqual(1, sheet.cell(9, columns["item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.value"]).value)
            result.close()


if __name__ == "__main__":
    unittest.main()
