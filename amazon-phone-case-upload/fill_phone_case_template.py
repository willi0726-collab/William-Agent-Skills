#!/usr/bin/env python3
"""Fill Amazon phone-case .xlsm templates and export upload-ready CRLF .txt."""

import os
import re
import openpyxl

INPUT_TEMPLATE = r"D:\path\to\downloaded_template.xlsm"
OUTPUT_DIR = r"D:\path\to\output"
BRAND = "FXFOOT"
PARENT_SKU = "CAOMEI-DJ-P"
CHILD_SKU_PREFIX = "CMDJ"
PRODUCT_TYPE = "cellularphonecase"
VARIATION_THEME = "SizeName"
MODELS = [("IP16", "iPhone 16", "iPhone 16"), ("IP17", "iPhone 17", "iPhone 17")]
COLORS = [("Clear", "Clear", "CLR", "")]

def make_title(size_name, color_word="", design_name="3D Strawberry"):
    color = f" {color_word}" if color_word else ""
    title = f"{BRAND} {design_name} Phone Case for {size_name}{color} Cute Soft TPU Clear"
    return title if len(title) <= 75 else f"{BRAND} {design_name} Case for {size_name}{color} Soft Clear TPU"

PARENT_TITLE = f"{BRAND} 3D Strawberry Phone Case Cute Soft TPU Clear Bumper"
ITEM_HIGHLIGHT = ""
BULLET1 = "3D strawberry design with a soft TPU protective case."
BULLET2 = "Flexible and lightweight for comfortable everyday use."
BULLET3 = "Raised camera bezel helps protect lenses from surface contact."
BULLET4 = "Precise cutouts provide access to buttons, speakers, and charging port."
BULLET5 = "Slim profile fits easily in pockets and bags."
DESCRIPTION = "A soft TPU phone case designed for everyday protection and precise fit."
SEARCH_TERMS = "strawberry phone case soft tpu protective cover"

PACKAGE_HEIGHT, PACKAGE_LENGTH, PACKAGE_WIDTH, PACKAGE_WEIGHT = 1.5, 12.0, 8.0, 60
PRICE, QUANTITY = 19.9, 0
MATERIAL, DESIGN_THEME, PATTERN, FORM_FACTOR = "Thermoplastic Polyurethane", "Floral", "Floral", "Basic Case"
INCLUDED_COMPONENTS = "Phone Case"
PRODUCT_ID_TYPE = "GTIN Exempt"
NUMBER_OF_ITEMS = 1
FINISH_TYPE = "Matte"
ITEM_THICKNESS = 0.15
ITEM_LENGTH = 12
ITEM_WIDTH = 5
ITEM_HEIGHT = 1
DIMENSION_UNIT = "Centimeters"
WEIGHT_UNIT = "Grams"
WARRANTY_DESCRIPTION = "1 Year Manufacture"
BATTERIES_REQUIRED = "No"
SKIP_OFFER = False
SHIPPING_TEMPLATE = ""
FULFILLMENT_CHANNEL = "Fulfillment by Merchant (Default)"

ALIASES = {
    "feed_product_type": ["product_type#1.value", "product_type"],
    "item_sku": ["contribution_sku#1.value", "item_sku"],
    "update_delete": ["::record_action", "record_action"],
    "brand_name": ["brand[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "parent_child": ["parentage_level[marketplace_id=ATVPDKIKX0DER]#1.value"],
    "parent_sku": ["child_parent_sku_relationship[marketplace_id=ATVPDKIKX0DER]#1.parent_sku"],
    "variation_theme": ["variation_theme#1.name"],
    "item_name": ["item_name[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "item_highlight": ["title_differentiation[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "manufacturer": ["manufacturer[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "product_description": ["product_description[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "external_product_id_type": ["amzn1.volt.ca.product_id_type"],
    "number_of_items": ["number_of_items[marketplace_id=ATVPDKIKX0DER]#1.value"],
    "part_number": ["part_number[marketplace_id=ATVPDKIKX0DER]#1.value"],
    "compatible_devices1": ["compatible_devices[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "finish_type": ["finish_type[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "thickness_decimal": ["item_thickness[marketplace_id=ATVPDKIKX0DER]#1.decimal_value"],
    "thickness_string": ["item_thickness[marketplace_id=ATVPDKIKX0DER]#1.string_value"],
    "thickness_unit": ["item_thickness[marketplace_id=ATVPDKIKX0DER]#1.unit"],
    "item_length": ["item_length[marketplace_id=ATVPDKIKX0DER]#1.value"],
    "item_length_unit": ["item_length[marketplace_id=ATVPDKIKX0DER]#1.unit"],
    "item_dimension_length": ["item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.value"],
    "item_dimension_length_unit": ["item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.unit"],
    "item_dimension_width": ["item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.value"],
    "item_dimension_width_unit": ["item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.unit"],
    "item_dimension_height": ["item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.value"],
    "item_dimension_height_unit": ["item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.unit"],
    "warranty": ["warranty_description[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "batteries_required": ["batteries_required[marketplace_id=ATVPDKIKX0DER]#1.value"],
    "bullet_point1": ["bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "bullet_point2": ["bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#2.value"],
    "bullet_point3": ["bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#3.value"],
    "bullet_point4": ["bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#4.value"],
    "bullet_point5": ["bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#5.value"],
    "generic_keywords": ["generic_keyword[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "special_features1": ["special_feature[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "size_name": ["size[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "color_name": ["color[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "material_type": ["material[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "case_material": ["case[marketplace_id=ATVPDKIKX0DER]#1.material[language_tag=en_US]#1.value"],
    "pattern_name": ["pattern[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "theme": ["theme[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "form_factor": ["form_factor[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "compatible_phone_models1": ["compatible_phone_models[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value"],
    "skip_offer": ["skip_offer[marketplace_id=ATVPDKIKX0DER]#1.value"],
    "condition_type": ["condition_type[marketplace_id=ATVPDKIKX0DER]#1.value"],
    "list_price": ["list_price[marketplace_id=ATVPDKIKX0DER]#1.value"],
    "fulfillment_channel": ["fulfillment_availability#1.fulfillment_channel_code"],
    "quantity": ["fulfillment_availability#1.quantity"],
    "price": ["purchasable_offer[marketplace_id=ATVPDKIKX0DER][audience=ALL]#1.our_price#1.schedule#1.value_with_tax", "purchasable_offer[marketplace_id=ATVPDKIKX0DER]#1.our_price#1.schedule#1.value_with_tax"],
    "shipping_template": ["merchant_shipping_group[marketplace_id=ATVPDKIKX0DER]#1.value"],
    "package_length": ["item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.value"],
    "package_length_unit_of_measure": ["item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.unit"],
    "package_width": ["item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.value"],
    "package_width_unit_of_measure": ["item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.unit"],
    "package_height": ["item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.value"],
    "package_height_unit_of_measure": ["item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.unit"],
    "package_weight": ["item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.value"],
    "package_weight_unit_of_measure": ["item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.unit"],
    "country_of_origin": ["country_of_origin[marketplace_id=ATVPDKIKX0DER]#1.value"],
}

def layout(ws):
    settings = str(ws.cell(1, 1).value or "")
    def number(key, default):
        match = re.search(rf"(?:^|&){key}=(\d+)", settings)
        return int(match.group(1)) if match else default
    return number("labelRow", 2), number("attributeRow", 3), number("dataRow", 4)

def columns(ws, attribute_row):
    raw = {str(ws.cell(attribute_row, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(attribute_row, c).value}
    mapped = dict(raw)
    for logical, candidates in ALIASES.items():
        for candidate in [logical] + candidates:
            if candidate in raw:
                mapped[logical] = raw[candidate]
                break
    return mapped

def put(ws, row, cols, field, value):
    if field in cols and value is not None and value != "":
        ws.cell(row, cols[field]).value = value

def precheck():
    errors = []
    for _, size, _ in MODELS:
        for _, _, _, color in COLORS:
            if len(make_title(size, color)) > 75: errors.append(f"Title over 75: {size}/{color}")
    for index, bullet in enumerate([BULLET1, BULLET2, BULLET3, BULLET4, BULLET5], 1):
        if len(bullet) > 125: errors.append(f"Bullet {index} over 125")
    if ITEM_HIGHLIGHT and len(ITEM_HIGHLIGHT) > 125: errors.append("Item Highlight over 125")
    if len(SEARCH_TERMS.encode("utf-8")) > 250: errors.append("Search terms over 250 bytes")
    if errors: raise ValueError("; ".join(errors))
    return True

def fill_template():
    precheck()
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(INPUT_TEMPLATE, keep_vba=True)
    ws = wb["Template"]
    _, attribute_row, data_row = layout(ws)
    cols = columns(ws, attribute_row)
    latest = attribute_row != 3 or "product_type#1.value" in cols
    for r in range(data_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1): ws.cell(r, c).value = None

    product_type = "CELLULAR_PHONE_CASE" if latest else PRODUCT_TYPE
    put(ws, data_row, cols, "feed_product_type", product_type)
    put(ws, data_row, cols, "item_sku", PARENT_SKU)
    put(ws, data_row, cols, "brand_name", BRAND)
    put(ws, data_row, cols, "parent_child", "Parent")
    put(ws, data_row, cols, "variation_theme", VARIATION_THEME)
    put(ws, data_row, cols, "item_name", PARENT_TITLE)
    put(ws, data_row, cols, "manufacturer", BRAND)
    put(ws, data_row, cols, "update_delete", "Create or Replace (Full Update)" if latest else None)
    put(ws, data_row, cols, "external_product_id_type", PRODUCT_ID_TYPE if latest else None)

    row = data_row + 1
    for color_name, color_map, color_suffix, color_word in COLORS:
        for sku_suffix, size_name, model_name in MODELS:
            sku = f"{CHILD_SKU_PREFIX}-{color_suffix}-{sku_suffix}" if len(COLORS) > 1 else f"{CHILD_SKU_PREFIX}-{sku_suffix}"
            values = {
                "feed_product_type": product_type, "item_sku": sku, "brand_name": BRAND,
                "update_delete": "Create or Replace (Full Update)" if latest else "Update", "item_name": make_title(size_name, color_word),
                "item_highlight": ITEM_HIGHLIGHT, "manufacturer": BRAND, "product_description": DESCRIPTION,
                "external_product_id_type": PRODUCT_ID_TYPE if latest else "GTIN Exemption", "parent_child": "Child",
                "parent_sku": PARENT_SKU, "relationship_type": "Variation", "variation_theme": VARIATION_THEME,
                "bullet_point1": BULLET1, "bullet_point2": BULLET2, "bullet_point3": BULLET3,
                "bullet_point4": BULLET4, "bullet_point5": BULLET5, "generic_keywords": SEARCH_TERMS,
                "special_features1": "Shock-Absorbent", "size_name": size_name, "size_map": size_name,
                "color_name": color_name, "color_map": color_map, "material_type": MATERIAL,
                "case_material": MATERIAL, "pattern_name": PATTERN, "theme": DESIGN_THEME,
                "form_factor": FORM_FACTOR, "included_components": INCLUDED_COMPONENTS,
                "compatible_phone_models1": model_name, "compatible_devices1": model_name,
                "number_of_items": NUMBER_OF_ITEMS, "part_number": sku, "finish_type": FINISH_TYPE,
                "thickness_decimal": ITEM_THICKNESS, "thickness_string": str(ITEM_THICKNESS),
                "thickness_unit": DIMENSION_UNIT, "warranty": WARRANTY_DESCRIPTION,
                "batteries_required": BATTERIES_REQUIRED,
                "item_length": ITEM_LENGTH, "item_length_unit": DIMENSION_UNIT,
                "item_dimension_length": ITEM_LENGTH, "item_dimension_length_unit": DIMENSION_UNIT,
                "item_dimension_width": ITEM_WIDTH, "item_dimension_width_unit": DIMENSION_UNIT,
                "item_dimension_height": ITEM_HEIGHT, "item_dimension_height_unit": DIMENSION_UNIT,
                "package_height": PACKAGE_HEIGHT,
                "package_length": PACKAGE_LENGTH, "package_width": PACKAGE_WIDTH, "package_weight": PACKAGE_WEIGHT,
                "package_height_unit_of_measure": DIMENSION_UNIT, "package_length_unit_of_measure": DIMENSION_UNIT,
                "package_width_unit_of_measure": DIMENSION_UNIT, "package_weight_unit_of_measure": WEIGHT_UNIT,
                "country_of_origin": "CN", "quantity": QUANTITY, "price": PRICE, "list_price": PRICE,
                "condition_type": "New", "fulfillment_channel": FULFILLMENT_CHANNEL,
                "shipping_template": SHIPPING_TEMPLATE, "skip_offer": "Yes" if SKIP_OFFER else None,
            }
            for field, value in values.items(): put(ws, row, cols, field, value)
            row += 1

    xlsm_path = os.path.join(OUTPUT_DIR, f"phone_case_{BRAND}_{PARENT_SKU}.xlsm")
    wb.save(xlsm_path)
    txt_path = os.path.join(OUTPUT_DIR, f"phone_case_{BRAND}_{PARENT_SKU}.txt")
    with open(txt_path, "w", encoding="utf-8", newline="") as handle:
        for r in range(1, row):
            values = [str(ws.cell(r, c).value or "").replace("\t", " ").replace("\r", " ").replace("\n", " ") for c in range(1, ws.max_column + 1)]
            handle.write("\t".join(values) + "\r\n")
    wb.close()
    return xlsm_path, txt_path

if __name__ == "__main__":
    fill_template()
